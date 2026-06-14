"""
Genera TAREAS_JOCHOA.xlsx — arquitectura v2 · Centro de Comando Personal
Migra 39 tareas con campos extendidos + hojas TERCEROS, REUNIONES, REFERENCIA.
"""
from datetime import date
import openpyxl as xl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HOY = date(2026, 6, 13)

# ── Estilos ──────────────────────────────────────────────────────────────────
BG_HDR  = "0D1526"
FG_HDR  = "38BDF8"
BG_ODD  = "0A1220"
BG_EVEN = "0D1830"
thin    = Side(style="thin", color="1E293B")
BRD     = Border(top=thin, bottom=thin, left=thin, right=thin)

def hdr(ws, row, col, val, w=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=True, color=FG_HDR, size=9, name="Segoe UI")
    c.fill      = PatternFill("solid", fgColor=BG_HDR)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = BRD
    if w: ws.column_dimensions[get_column_letter(col)].width = w
    return c

def dat(ws, row, col, val, odd=True, clr="CBD5E1", center=False, wrap=False, date_fmt=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(color=clr, size=9, name="Segoe UI")
    c.fill      = PatternFill("solid", fgColor=BG_ODD if odd else BG_EVEN)
    c.border    = BRD
    c.alignment = Alignment(vertical="center", horizontal="center" if center else "left",
                            wrap_text=wrap or (isinstance(val, str) and len(str(val or "")) > 35))
    if date_fmt and isinstance(val, date):
        c.number_format = "DD/MM/YYYY"
    return c

# ═══════════════════════════════════════════════════════════════════════════════
# HOJA 1: TAREAS
# Columnas: ID | TAREA | DESCRIPCION | TIPO | PROYECTO | AREA | CATEGORIA |
#           PRIORIDAD | ESTADO | IMPACTO | URGENCIA | ESFUERZO_HRS | TERCERO |
#           FECHA_CREACION | FECHA_COMPROMISO | FECHA_CIERRE | NOTAS
# IMPACTO y URGENCIA: escala 1–5
# ESFUERZO_HRS: horas estimadas
# ═══════════════════════════════════════════════════════════════════════════════

TAREAS_HDRS = [
    ("ID",              4),   ("TAREA",          44),  ("DESCRIPCION",    30),
    ("TIPO",           12),   ("PROYECTO",       14),  ("AREA",           10),
    ("CATEGORIA",      20),   ("PRIORIDAD",      10),  ("ESTADO",         20),
    ("IMPACTO",         8),   ("URGENCIA",        8),  ("ESFUERZO_HRS",   12),
    ("TERCERO",        18),   ("FECHA_CREACION", 14),  ("FECHA_COMPROMISO",14),
    ("FECHA_CIERRE",   14),   ("NOTAS",          30),
]

# Tupla: (ID, TAREA, DESCRIPCION, TIPO, PROYECTO, AREA, CATEGORIA,
#         PRIORIDAD, ESTADO, IMPACTO, URGENCIA, ESFUERZO_HRS, TERCERO,
#         FECHA_CREACION, FECHA_COMPROMISO, FECHA_CIERRE, NOTAS)
TAREAS = [
    (1,  "REVISIÓN plan de abastecimiento CL038",
         "", "Tarea", "CL038", "Trabajo", "Planificación",
         "Media", "Pendiente", 3, 4, 1.5, "",
         date(2026,6,7), date(2026,6,19), None, ""),

    (2,  "SEGUIMIENTO correo clínica MEDs (carpeta rendición Isapre)",
         "", "Seguimiento", "SALUD", "Personal", "Salud",
         "Media", "Completada", 3, 5, 0.5, "Clínica MEDs",
         date(2026,6,6), date(2026,6,13), date(2026,6,13), ""),

    (3,  "LLEVAR formulario seguro (boletas terapía)",
         "", "Tarea", "SALUD", "Personal", "Salud",
         "Media", "Pendiente", 3, 4, 0.5, "",
         date(2026,6,10), date(2026,6,16), None, ""),

    (4,  "SEGUIMIENTO IConstruye (reportabilidad gastado a la fecha)",
         "", "Seguimiento", "ICONSTRUYE", "Trabajo", "Reportes",
         "Media", "Pendiente", 3, 5, 1.0, "IConstruye",
         date(2026,6,8), date(2026,6,15), None, ""),

    (5,  "APROBAR contratos",
         "", "Compromiso", "CORPORATIVO", "Trabajo", "Contratos",
         "Media", "Pendiente", 3, 5, 1.0, "",
         date(2026,6,10), date(2026,6,14), None, ""),

    (6,  "Aprobar regularización equipos robados",
         "", "Compromiso", "CORPORATIVO", "Trabajo", "Gestión Corporativa",
         "Media", "Pendiente", 3, 4, 1.0, "",
         date(2026,6,10), date(2026,6,16), None, ""),

    (7,  "VISITAR a mi tío",
         "", "Tarea", "PERSONAL", "Personal", "Personal",
         "Media", "Pendiente", 2, 5, 2.0, "",
         date(2026,6,12), date(2026,6,14), None, ""),

    (8,  "PASAR aspiradora",
         "", "Tarea", "PERSONAL", "Personal", "Personal",
         "Media", "Pendiente", 1, 5, 0.5, "",
         date(2026,6,12), date(2026,6,14), None, ""),

    (9,  "HACER baño 1",
         "", "Tarea", "PERSONAL", "Personal", "Personal",
         "Media", "Pendiente", 1, 5, 0.5, "",
         date(2026,6,12), date(2026,6,14), None, ""),

    (10, "ENVIAR correo presentación directorio (diferido)",
         "", "Compromiso", "DIRECTORIO", "Trabajo", "Gestión Corporativa",
         "Media", "Pendiente", 4, 5, 0.5, "Directorio",
         date(2026,6,10), date(2026,6,14), None, ""),

    (11, "SOLICITAR a GB respaldos de acta de sesión de Directorios",
         "", "Seguimiento", "DIRECTORIO", "Trabajo", "Gestión Corporativa",
         "Media", "En Proceso", 4, 4, 0.5, "GB",
         date(2026,6,10), date(2026,6,17), None, ""),

    (12, "ACTUALIZAR formato control de hormigones CL038",
         "", "Tarea", "CL038", "Trabajo", "Planificación",
         "Media", "Pendiente", 3, 5, 2.0, "",
         date(2026,6,9), date(2026,6,14), None, ""),

    (13, "RESERVA hora con Javi Franco",
         "", "Tarea", "SALUD", "Personal", "Salud",
         "Media", "En Proceso", 2, 5, 0.5, "Javi Franco",
         date(2026,6,12), date(2026,6,13), None, ""),

    (14, "GESTIONAR firma MPD",
         "", "Compromiso", "CORPORATIVO", "Trabajo", "Gestión Corporativa",
         "Media", "En Proceso", 4, 4, 1.5, "MPD",
         date(2026,6,9), date(2026,6,16), None, ""),

    (15, "GESTIONAR firma Contrato Salfa-Londres",
         "", "Compromiso", "CORPORATIVO", "Trabajo", "Contratos",
         "Media", "En Proceso", 5, 4, 1.5, "Salfa",
         date(2026,6,9), date(2026,6,16), None, ""),

    (16, "ENVIAR comunicado por día de vacaciones",
         "", "Compromiso", "CORPORATIVO", "Personal", "Gestión Corporativa",
         "Alta", "En Proceso", 4, 4, 0.5, "",
         date(2026,6,10), date(2026,6,16), None, ""),

    (17, "PRESENTAR Dashboard Control de áridos",
         "", "Compromiso", "AGENTE_IA", "Trabajo", "IA y Automatización",
         "Media", "Pendiente", 4, 4, 2.0, "",
         date(2026,6,10), date(2026,6,16), None, ""),

    (18, "LAVAR ropa",
         "", "Tarea", "PERSONAL", "Personal", "Personal",
         "Media", "Pendiente", 1, 5, 0.5, "",
         date(2026,6,12), date(2026,6,13), None, ""),

    (19, "LAVAR losa",
         "", "Tarea", "PERSONAL", "Personal", "Personal",
         "Media", "Pendiente", 1, 5, 0.5, "",
         date(2026,6,12), date(2026,6,13), None, ""),

    (20, "CAPACITACIÓN Claude Code 3/5",
         "", "Tarea", "AGENTE_IA", "Trabajo", "IA y Automatización",
         "Media", "Pendiente", 3, 5, 2.0, "",
         date(2026,6,8), date(2026,6,14), None, ""),

    (21, "CAPACITACIÓN Claude Code 4/5",
         "", "Tarea", "AGENTE_IA", "Trabajo", "IA y Automatización",
         "Media", "Pendiente", 3, 5, 2.0, "",
         date(2026,6,8), date(2026,6,15), None, ""),

    (22, "CAPACITACIÓN Claude Code 5/5",
         "", "Tarea", "AGENTE_IA", "Trabajo", "IA y Automatización",
         "Media", "Pendiente", 3, 4, 2.0, "",
         date(2026,6,8), date(2026,6,16), None, ""),

    (23, "BARRER hojas",
         "", "Tarea", "PERSONAL", "Personal", "Personal",
         "Media", "Pendiente", 1, 5, 0.5, "",
         date(2026,6,12), date(2026,6,13), None, ""),

    (24, "PREPARAR almuerzo",
         "", "Tarea", "PERSONAL", "Personal", "Personal",
         "Media", "Pendiente", 1, 5, 0.5, "",
         date(2026,6,12), date(2026,6,13), None, ""),

    (25, "CAPACITACIÓN Claude Code 2/5",
         "", "Tarea", "AGENTE_IA", "Trabajo", "IA y Automatización",
         "Media", "Pendiente", 3, 5, 2.0, "",
         date(2026,6,8), date(2026,6,13), None, ""),

    (26, "LLAMAR a Claudio S.",
         "", "Tarea", "PERSONAL", "Personal", "Personal",
         "Alta", "Pendiente", 3, 5, 0.5, "Claudio S.",
         date(2026,6,12), date(2026,6,13), None, ""),

    (27, "ASISTIR cumpleaños Felipe M.",
         "", "Tarea", "PERSONAL", "Personal", "Personal",
         "Media", "Pendiente", 2, 5, 2.0, "",
         date(2026,6,12), date(2026,6,13), None, ""),

    (28, "ASISTIR cumpleaños Esteban",
         "", "Tarea", "PERSONAL", "Personal", "Personal",
         "Media", "Pendiente", 2, 5, 2.0, "",
         date(2026,6,12), date(2026,6,13), None, ""),

    (29, "ENVIAR formato control de hormigones CL038",
         "", "Compromiso", "CL038", "Trabajo", "Planificación",
         "Media", "En Proceso", 3, 4, 0.5, "",
         date(2026,6,9), date(2026,6,16), None, ""),

    (30, "PRESENTAR carpetas 2 ciclos fertilización en isapre",
         "", "Compromiso", "SALUD", "Personal", "Salud",
         "Alta", "Pendiente", 4, 3, 2.0, "Isapre",
         date(2026,6,9), date(2026,6,20), None, ""),

    (31, "REVISIÓN funcionamiento KIT en IConstruye",
         "", "Seguimiento", "ICONSTRUYE", "Trabajo", "Reportes",
         "Media", "Pendiente", 3, 3, 1.5, "IConstruye",
         date(2026,6,9), date(2026,6,18), None, ""),

    (32, "ACTUALIZAR modelo Claude (validar datos)",
         "", "Tarea", "AGENTE_IA", "Trabajo", "IA y Automatización",
         "Media", "En Proceso", 3, 5, 2.0, "",
         date(2026,6,10), date(2026,6,14), None, ""),

    (33, "COORDINAR hora psiquiatra (Consultar con la Lore)",
         "", "Tarea", "SALUD", "Personal", "Salud",
         "Media", "Pendiente", 3, 5, 0.5, "Consultorio",
         date(2026,6,12), date(2026,6,13), None, ""),

    (34, "COMPRAR nuevo anillo matrimonial",
         "", "Tarea", "PERSONAL", "Personal", "Personal",
         "Baja", "Pendiente", 2, 3, 1.0, "",
         date(2026,6,10), date(2026,6,20), None, ""),

    (35, "ENVIAR a reparar bicicleta",
         "", "Tarea", "PERSONAL", "Personal", "Personal",
         "Media", "Pendiente", 2, 3, 0.5, "",
         date(2026,6,10), date(2026,6,20), None, ""),

    (36, "ENTRENAR trote: Rodaje 5K (Matutino)",
         "", "Tarea", "PERSONAL", "Personal", "Personal",
         "Alta", "Pendiente", 3, 5, 1.0, "",
         date(2026,6,12), date(2026,6,15), None, ""),

    (37, "ARMAR propuesta para salida fin de mes",
         "", "Tarea", "PERSONAL", "Personal", "Personal",
         "Alta", "Pendiente", 3, 5, 1.5, "",
         date(2026,6,12), date(2026,6,14), None, ""),

    (38, "Revisar propuestas PUC RSC",
         "", "Tarea", "CORPORATIVO", "Trabajo", "Gestión Corporativa",
         "Media", "Pendiente", 3, 1, 1.0, "",
         date(2026,6,9), None, None, ""),

    (39, "Consolidar ensayos Mov. Tierra CL027",
         "", "Tarea", "CL027", "Trabajo", "Planificación",
         "Media", "Pendiente", 3, 1, 2.0, "",
         date(2026,6,9), None, None, ""),
]

# ═══════════════════════════════════════════════════════════════════════════════
# HOJA 2: TERCEROS
# ID | NOMBRE | ORGANIZACION | ROL | TEMA_PENDIENTE |
# FECHA_INICIO_SEG | FECHA_ULTIMO_SEG | ESTADO | PRIORIDAD | NOTAS
# ═══════════════════════════════════════════════════════════════════════════════
TERC_HDRS = [
    ("ID",6),("NOMBRE",18),("ORGANIZACION",20),("ROL",18),
    ("TEMA_PENDIENTE",35),("FECHA_INICIO_SEG",16),("FECHA_ULTIMO_SEG",16),
    ("ESTADO",14),("PRIORIDAD",10),("NOTAS",30),
]
TERCEROS = [
    (1, "IConstruye", "Plataforma Digital", "Proveedor de Servicio",
     "Funcionamiento KIT y reporte reportabilidad",
     date(2026,6,9), date(2026,6,9), "Pendiente", "Alta", "Abrir ticket si no responden en 48h"),

    (2, "GB", "Gerencia Buenos Aires", "Gerencia Corporativa",
     "Respaldos actas de sesión Directorio",
     date(2026,6,10), date(2026,6,11), "Pendiente", "Media", "Recordar para reunión directorio"),

    (3, "MPD", "Mandante Proyecto", "Mandante",
     "Firma contrato pendiente",
     date(2026,6,9), date(2026,6,9), "Pendiente", "Alta", "Bloquea inicio de faena"),

    (4, "Salfa", "Salfa Contratistas", "Contratista",
     "Firma Contrato Salfa-Londres",
     date(2026,6,9), date(2026,6,9), "Pendiente", "Alta", "Coordinar con legal si hay demora"),

    (5, "Isapre", "Institución de Salud", "Institución Salud",
     "Presentación carpetas 2 ciclos fertilización",
     date(2026,6,6), date(2026,6,6), "Pendiente", "Alta", "Plazo vence 20 junio"),

    (6, "Clínica MEDs", "Clínica Médica", "Prestador Salud",
     "Correo rendición Isapre",
     date(2026,6,6), date(2026,6,13), "Resuelto", "Baja", "Correo recibido"),
]

# ═══════════════════════════════════════════════════════════════════════════════
# HOJA 3: REUNIONES
# ID | TITULO | FECHA | TIPO | PARTICIPANTES | ESTADO |
# COMPROMISO | RESPONSABLE_COMP | FECHA_COMP | ESTADO_COMP | NOTAS
# ═══════════════════════════════════════════════════════════════════════════════
REUN_HDRS = [
    ("ID",5),("TITULO",28),("FECHA",14),("TIPO",16),
    ("PARTICIPANTES",28),("ESTADO",12),("COMPROMISO",35),
    ("RESPONSABLE_COMP",18),("FECHA_COMP",14),("ESTADO_COMP",14),("NOTAS",25),
]
REUNIONES = [
    (1, "Reunión Directorio — Estado de Obras", date(2026,6,20),
     "Directorio", "Gerencia, Directores, ITO", "Pendiente",
     "Preparar presentación estado obras; Enviar correo previo con agenda",
     "Juan", date(2026,6,18), "En Proceso",
     "Presentar avance CL038 y CL027"),

    (2, "Coordinación Avance CL038", date(2026,6,17),
     "Obra", "ITO, Administrador de Obra, Subcontratistas", "Pendiente",
     "Enviar informe de avance; Actualizar formato hormigones",
     "Juan", date(2026,6,16), "Pendiente",
     "Revisar plan de abastecimiento áridos"),

    (3, "Revisión Contratos — Salfa / MPD", date(2026,6,16),
     "Contratos", "Gerencia, Asesor Legal, Representante Salfa", "Pendiente",
     "Gestionar firma MPD; Gestionar firma Salfa; Enviar copia a legal",
     "Juan", date(2026,6,15), "En Proceso",
     "Confirmar asistencia de ambas partes"),

    (4, "Seguimiento IConstruye — Funcionalidades", date(2026,6,18),
     "Tecnología", "Soporte IConstruye, Planificación", "Pendiente",
     "Resolver funcionamiento KIT; Revisar reportabilidad gastado",
     "Juan", date(2026,6,18), "Pendiente",
     "Preparar pantallazos con el error"),
]

# ═══════════════════════════════════════════════════════════════════════════════
# HOJA 4: REFERENCIA
# ═══════════════════════════════════════════════════════════════════════════════
REFERENCIA = {
    "PROYECTO":   ["CL038","CL027","ICONSTRUYE","AGENTE_IA","CORPORATIVO",
                   "DIRECTORIO","SALUD","PERSONAL","GENERAL"],
    "PRIORIDAD":  ["Crítica","Alta","Media","Baja"],
    "ESTADO":     ["Pendiente","En Proceso","Esperando Terceros","Completada","Cancelada"],
    "AREA":       ["Trabajo","Personal"],
    "TIPO":       ["Tarea","Seguimiento","Compromiso","Reunión"],
    "CATEGORIA":  ["Planificación","Contratos","Compras","Reportes",
                   "IA y Automatización","Gestión Corporativa","Reuniones","Salud","Personal"],
    "IMPACTO":    ["1 - Mínimo","2 - Bajo","3 - Medio","4 - Alto","5 - Crítico"],
    "URGENCIA":   ["1 - Sin urgencia","2 - Baja","3 - Media","4 - Alta","5 - Inmediata"],
}

# ─── CONSTRUIR WORKBOOK ──────────────────────────────────────────────────────
wb = xl.Workbook()
wb.remove(wb.active)

# ── Hoja TAREAS ──────────────────────────────────────────────────────────────
ws1 = wb.create_sheet("TAREAS")
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A2"
ws1.row_dimensions[1].height = 28

for ci, (h, w) in enumerate(TAREAS_HDRS, 1):
    hdr(ws1, 1, ci, h, w)

for ri, t in enumerate(TAREAS, 2):
    odd = (ri % 2 == 0)
    for ci, val in enumerate(t, 1):
        is_date = isinstance(val, date)
        clr = "94A3B8" if ci in (3,13,17) else "CBD5E1"
        dat(ws1, ri, ci, val, odd, clr,
            center=(ci in (1,4,6,7,8,9,10,11,12)),
            date_fmt=is_date)
    ws1.row_dimensions[ri].height = 20

# ── Hoja TERCEROS ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("TERCEROS")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A2"
ws2.row_dimensions[1].height = 28

for ci, (h, w) in enumerate(TERC_HDRS, 1):
    hdr(ws2, 1, ci, h, w)

for ri, t in enumerate(TERCEROS, 2):
    odd = (ri % 2 == 0)
    for ci, val in enumerate(t, 1):
        dat(ws2, ri, ci, val, odd,
            center=(ci in (1,6,7,8,9)),
            date_fmt=isinstance(val, date))
    ws2.row_dimensions[ri].height = 20

# ── Hoja REUNIONES ────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("REUNIONES")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A2"
ws3.row_dimensions[1].height = 28

for ci, (h, w) in enumerate(REUN_HDRS, 1):
    hdr(ws3, 1, ci, h, w)

for ri, t in enumerate(REUNIONES, 2):
    odd = (ri % 2 == 0)
    for ci, val in enumerate(t, 1):
        dat(ws3, ri, ci, val, odd,
            center=(ci in (1,3,4,6,8,9,10)),
            date_fmt=isinstance(val, date))
    ws3.row_dimensions[ri].height = 22

# ── Hoja REFERENCIA ───────────────────────────────────────────────────────────
ws4 = wb.create_sheet("REFERENCIA")
ws4.sheet_view.showGridLines = False
col_i = 1
for h, vals in REFERENCIA.items():
    hdr(ws4, 1, col_i, h, 22)
    ws4.row_dimensions[1].height = 22
    for ri, v in enumerate(vals, 2):
        c = ws4.cell(ri, col_i, v)
        c.font = Font(color="94A3B8", size=9, name="Segoe UI")
        c.fill = PatternFill("solid", fgColor=BG_ODD if ri % 2 == 0 else BG_EVEN)
        ws4.row_dimensions[ri].height = 18
    col_i += 1

OUT = str(__file__).replace("crear_excel_v2.py", "TAREAS_JOCHOA.xlsx")
wb.save(OUT)
n = len(TAREAS)
print(f"Excel v2 creado: {OUT}")
print(f"   {n} tareas, {len(TERCEROS)} terceros, {len(REUNIONES)} reuniones")
